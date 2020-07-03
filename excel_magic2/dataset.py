import datetime
from collections.abc import MutableMapping
from copy import copy
import sqlite3
from io import BytesIO
from typing import Callable, Union, List, Any, Tuple, Dict
import os
import shutil
import xlsxwriter
import csv
import json
from PIL import Image
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Worksheet
import openpyxl.cell
from excel_magic2.Exceptions import EmptySheetException, PredictionException

__all__ = ['Dataset', 'open_file']


class Pointer:
    def __init__(self, row: int, col: int):
        self.row = row
        self.col = col

    def next_row(self, current_col=False):
        if not current_col:
            self.col = 0
        self.row += 1

    def next_col(self):
        self.col += 1


class HorizontalAlignment:
    LEFT = 'left'
    CENTER = 'center'
    RIGHT = 'right'


class VerticalAlignment:
    TOP = 'top'
    CENTER = 'center'
    BOTTOM = 'bottom'


class Style:
    def __init__(self, horizontal_alignment='left',
                 vertical_alignment='top',
                 bold=False,
                 underline=False,
                 font_color='black',
                 font_name='Calibri',
                 font_size='12',
                 fill_color=''):
        self.horizontal_alignment = horizontal_alignment
        self.vertical_alignment = vertical_alignment
        self.bold = bold
        self.underline = underline
        self.font_color = font_color
        self.font_name = font_name
        self.font_size = font_size
        self.fill_color = fill_color
        self.num_format = ''

    def __copy__(self):
        result = Style(self.horizontal_alignment,
                       self.vertical_alignment,
                       self.bold,
                       self.underline,
                       self.font_color,
                       self.font_name,
                       self.font_size,
                       self.fill_color)
        result.num_format = self.num_format
        return result

    def attr(self):
        attr = {'align'     : self.horizontal_alignment,
                'valign'    : self.vertical_alignment,
                'bold'      : self.bold,
                'underline' : self.underline,
                'font_color': self.font_color,
                'font_name' : self.font_name,
                'font_size' : self.font_size}
        if self.fill_color != '':
            attr['bg_color'] = self.fill_color
        if self.num_format != '':
            attr['num_format'] = self.num_format
        return attr


class Header:
    def __init__(self, value: str, style: Style, width: int = 20):
        self.value = value
        self.style = style
        self.width = width


class Cell:
    def __init__(self, value: Any = '', style: Style = None):
        self._value = value
        if style is None:
            self.style = Style()
        else:
            self.style = style

    @property
    def value(self):
        if isinstance(self._value, float) and self._value % 1 == 0:
            return int(self._value)
        else:
            return self._value

    @value.setter
    def value(self, value):
        self._value = value

    def __copy__(self):
        return Cell(self.value)

    def set_style(self, style: Style):
        self.style = style

    def attr(self):
        return self.style.attr()

    def __str__(self):
        return str(self.value)

    def __eq__(self, other: Union['Cell', str]):
        if isinstance(other, str):
            return self.value == other
        elif isinstance(other, Cell):
            return self.value == other.value and \
                   self.style == other.style
        else:
            if isinstance(self.value, type(other)):
                return self.value == other
            else:
                if isinstance(other, int):
                    return self.value == float(other)
                return self.value is other


class ImageCell(Cell):
    def __init__(self, data: Union[BytesIO, str]):
        super().__init__()
        self.data = data
        self.value = ''

    def __copy__(self):
        return ImageCell(self.data)


class FormulaCell(Cell):
    def __init__(self, value: Any = '', formula: str = '', style: Style = None):
        super().__init__()
        self.formula = formula

    def __copy__(self):
        return FormulaCell(formula=self.formula)


class MergedCell(Cell):
    def __init__(self, start_pos: Tuple[int, int] = None, end_pos: Tuple[int, int] = None, value: Any = '', style: Style = None):
        """
        represents a merged cell
        :param start_pos: Tuple[row, col]
        :param end_pos: Tuple[row, col]
        :param value:
        :param style:
        """
        super().__init__(value, style)
        self.start_row, self.start_col = start_pos if start_pos is not None else (-1,-1)
        self.end_row, self.end_col = end_pos if end_pos is not None else (-1,-1)

    def add_col(self):
        self.end_col += 1

    def add_row(self):
        self.end_row += 1


class Row(MutableMapping):
    def __init__(self, fields: List[str]):
        self.fields = fields
        self.raw: Dict[Cell] = {}

    def __getitem__(self, item):
        return self.raw[item]

    def __setitem__(self, key, value):
        if isinstance(value, Cell):
            self.raw[key] = value
        else:
            self.raw[key] = Cell(value)

    def __iter__(self):
        return self.fields.__iter__()

    def __delitem__(self, key):
        del self.raw[key]

    def __len__(self):
        return len(self.raw)

    def __contains__(self, item):
        return item in self.raw

    def __copy__(self):
        result = Row(self.fields)
        for i in self.raw:
            result[i] = copy(self.raw[i])
        return result

    def filter_fields(self, cols: List[str]) -> 'Row':
        row = Row([])
        for col in self.fields:
            if col in cols:
                row.fields.append(col)
        for col in row.fields:
            row[col] = copy(self.raw[col])
        return row

    def __str__(self):
        result = '{'
        for i in self.raw:
            result += f'"{i}": {self.raw[i].value}; '
        result += '}'
        return result

    def __repr__(self):
        return self.__str__()

    def __eq__(self, other):
        if isinstance(other, Row):
            if self.fields == other.fields:
                for i in self.fields:
                    if not ((i in self.raw and i in other.raw) or (i not in self.raw and i not in other.raw)):
                        return False
                    else:
                        if self.raw[i].value != other.raw[i].value:
                            return False
                else:
                    return True
            else:
                return False
        else:
            return False

    def _intersect(self, b: 'Row'):
        result = Row([])

        for i in self.raw:
            if i in b.raw:
                result.fields.append(i)
                result[i] = self[i]
        return result

    def _union(self, b: 'Row'):
        result = Row([])
        for i in self.raw:
            result[i] = copy(self.raw[i])
        for i in b.raw:
            if i not in result.raw:
                result[i] = copy(b.raw[i])

        return result

    def __add__(self, other: Union['Row', Dict[str, str]]) -> 'Row':
        result = Row([])

        for col in self:
            result.fields.append(col)
            result[col] = copy(self[col])

        for col in other:
            if col in self and self[col].value != (other[col].value if isinstance(other, Row) else other[col]):
                raise ValueError('Unable to add two row having the same header but different values')
            result.fields.append(col)
            result[col] = copy(other[col])

        return result

    def __sub__(self, other: Union['Row', Dict[str, str]]) -> 'Row':
        result = Row([])

        for col in self:
            if col not in other:
                result[col] = copy(self[col])

        return result

    def values(self):
        return self.raw.values()

    def keys(self):
        return self.fields


class Sheet:
    def __init__(self, suppress_warning: bool = False, sheet: Union[Worksheet, str] = ''):
        self.fields = []
        self.data_rows: List[Row] = []
        self.header_style: Style = Style()
        self.suppress_warning = suppress_warning
        self.header_starts = -1
        self.file_head: List[List[Cell]] = []
        self.raw_sheet = sheet
        if isinstance(sheet, str):
            self.name: str = sheet
        else:
            self.name: str = sheet.title
            self._predict_header_pos(sheet)
            self._init_fields(sheet)
            self._init_data(sheet)

    def __len__(self):
        return self.data_rows.__len__()

    def sheet_length(self):
        return self.__len__()

    def _predict_header_pos(self, sheet: Worksheet):
        same_cell_count_run = 0
        last_cell_count = 0
        row_counter = 0
        possible_header = None
        header_starts = -1
        head_rows = []
        for row in sheet.rows:
            current_merged_cell = None
            head_row = []
            col_counter = 0
            has_merged = False
            for cell in row:
                # if it is a normal cell
                if isinstance(cell, openpyxl.cell.Cell):
                    # look forward
                    # if it is not the last cell
                    if col_counter != row.__len__() -1:
                        if isinstance(row[col_counter + 1], openpyxl.cell.MergedCell):
                            # look right
                            # If the next cell is a merged cell, this is the starting point of a merged cell(horizontal)
                            current_merged_cell = MergedCell((row_counter, col_counter), value=cell.value)
                            head_row.append(current_merged_cell)
                            has_merged = True

                        elif isinstance(sheet.cell(row_counter + 2, col_counter + 1), openpyxl.cell.MergedCell):
                            # look down
                            # if it is a merged cell, this is a merged cell(vertical)
                            current_merged_cell = MergedCell((row_counter, col_counter),
                                                             value=sheet.cell(row_counter + 1, col_counter + 1).value)
                            head_row.append(current_merged_cell)
                            has_merged = True

                        else:
                            # if there is a merged cell this row, it ends at the lefter col
                            if current_merged_cell is not None:
                                current_merged_cell.end_col = col_counter - 1
                            # look up, if there is a merged cell, it ends at the upper row
                            if len(head_rows) != 0:
                                if isinstance(head_rows[row_counter - 1][col_counter], MergedCell):
                                    head_rows[row_counter - 1][col_counter].end_row = row_counter - 1
                                head_row.append(Cell(cell.value))
                    else:
                        if current_merged_cell is not None:
                            current_merged_cell.end_col = col_counter - 1
                        # look up, if there is a merged cell, it ends at the upper row
                        if len(head_rows) != 0:
                            if isinstance(head_rows[row_counter - 1][col_counter], MergedCell):
                                head_rows[row_counter - 1][col_counter].end_row = row_counter - 1
                            head_row.append(Cell(cell.value))

                # if it is a merged cell
                elif isinstance(cell, openpyxl.cell.MergedCell):
                    has_merged = True
                    # find the correct merged cell
                    # look forward if it is not the first cell of this row
                    if col_counter != 0:
                        # if the lefter one is a merged cell and it has not ended yet
                        if isinstance(head_row[col_counter - 1], MergedCell) and head_row[col_counter - 1].end_col == -1:
                            # it belongs to the lefter merged cell
                            head_row.append(head_row[col_counter - 1])
                            continue
                    # look up if it is not the first row
                    if row_counter != 0:
                        # if it is a merged cell and it has not ended yet
                        if isinstance(head_rows[row_counter - 1][col_counter], MergedCell) and\
                                head_rows[row_counter - 1][col_counter].end_col == -1:
                            # it belongs to the upper merged cell
                            head_row.append(head_rows[row_counter - 1][col_counter])
                        else:
                            raise PredictionException('Unable to find the corresponding starting merged cell')
                    else:
                        raise PredictionException('Unable to find the corresponding starting merged cell')

                col_counter += 1

            head_rows.append(head_row)
            if not has_merged:
                # may be header
                if last_cell_count == head_row.__len__():
                    same_cell_count_run += 1
                last_cell_count = head_row.__len__()
                if possible_header is None:
                    possible_header = head_row
                    header_starts = row_counter
            else:
                last_cell_count = 0
                possible_header = None
                header_starts = -1

            # if there are more than 2
            if same_cell_count_run > 2:
                self.header_starts = header_starts
                self.file_head = head_rows
                break

            row_counter += 1
        else:
            if same_cell_count_run > 0:
                # assume it is header
                self.header_starts = header_starts
                self.file_head = head_rows
            else:
                raise PredictionException('Unable to find header')

    def _init_fields(self, sheet: Worksheet):
        counter = 0
        fields_row = None
        for row in sheet.rows:
            # skip headers
            if counter < self.header_starts:
                counter += 1
                continue
            fields_row = row
            break
        for field in fields_row:
            self.fields.append(field.value)

    def _init_data(self, sheet: Worksheet):
        row: Tuple[openpyxl.cell.Cell]
        counter = 0
        for row in sheet.rows:
            # skip headers
            if counter < self.header_starts + 1:
                counter += 1
                continue

            new_row = Row(self.fields)
            for i in range(len(self.fields)):
                # to prevent bug when there is an empty cell
                if i < len(row):
                    if isinstance(row[i].value, datetime.datetime):
                        value: datetime.datetime = row[i].value
                        if isinstance(row[i].value, datetime.time):
                            c = Cell(value)
                        else:
                            if value.hour == 0 and value.minute == 0 and value.second == 0:
                                c = Cell(value.date())
                            else:
                                c = Cell(value)
                        new_row[self.fields[i]] = c
                    else:
                        if isinstance(row[i].value, str):
                            try:
                                if row[i].value.isascii() and row[i].value.isnumeric():
                                    if not self.suppress_warning:
                                        print('Warning: Found a number stored in string format, converting...')
                                    new_row[self.fields[i]] = Cell(float(row[i].value))
                            except AttributeError:
                                if not self.suppress_warning:
                                    print('Warning: python3.6 compatibility mode')
                        new_row[self.fields[i]] = Cell(row[i].value)
                else:
                    new_row[self.fields[i]] = ''
            self.data_rows.append(new_row)
        for cell in self.data_rows[-1].values():
            if cell.value is not None:
                break
        else:
            self.data_rows.pop(-1)

    def set_header_style(self, style: Style):
        self.header_style = style

    def duplicate(self, name: str, headers_only: bool = False):
        result = Sheet(sheet=name)
        result.fields = [*self.fields]
        if not headers_only:
            for row in self.data_rows:
                r = {**row}
                result.append_row(r)
        return result

    def find(self, pairs: Union[dict, None] = None, none_if_not_found=False, **kwargs) -> Union[List[Row], None]:
        result = []
        if pairs is not None:
            kwargs = pairs
        # Check kwargs
        for kwarg in kwargs:
            if kwarg not in self.fields:
                raise NameError(f'field {kwarg} not found')

        for data_row in self.data_rows:
            for key in kwargs.keys():
                if isinstance(kwargs[key], int):
                    if data_row[key].value != float(kwargs[key]):
                        break
                    else:
                        continue

                if isinstance(kwargs[key], Cell):
                    if data_row[key].value != kwargs[key].value:
                        break
                    else:
                        continue

                if data_row[key].value != kwargs[key]:
                    break
            else:
                result.append(data_row)
        if result.__len__() == 0 and none_if_not_found:
            return None
        return result

    def highlight(self, rows: List[Row], highlight_style: Style):
        for row in rows:
            result = self.find(**row)
            for r in result:
                self.set_row_style(r, highlight_style)

    def filter(self, callback: Callable[[Row], Union[None, bool]]) -> List[Row]:
        data_list = []

        for row in self.data_rows:
            result = callback(row)
            if bool(result):
                data_list.append(row)

        return data_list

    def append_row(self, content: Union[Row, dict, List[Union[str, Cell]]]) -> None:
        new_row = Row(self.fields)
        if isinstance(content, dict) or isinstance(content, Row):
            for field in self.fields:
                if field in content:
                    if isinstance(content[field], Cell):
                        new_row[field] = content[field]
                    else:
                        new_row[field] = Cell(content[field])
                else:
                    new_row[field] = Cell('')
        elif isinstance(content, list):
            for i in range(len(self.fields)):
                if isinstance(content[i], Cell):
                    new_row[self.fields[i]] = content[i]
                else:
                    new_row[self.fields[i]] = Cell(content[i])

            if len(content) < len(self.fields):
                for i in range(len(content) - 1, len(self.fields)):
                    new_row[self.fields[i]] = Cell('')
        else:
            raise TypeError('Expected Row, dict or list')
        self.data_rows.append(new_row)

    def append_rows(self, rows: List[Union[dict, Row, List]]):
        for row in rows:
            self.append_row(row)

    def get_rows(self) -> List[Row]:
        r = [*self.data_rows]
        return r

    def get_col(self, col: str):
        if col not in self.fields:
            raise NameError(f'field "{col}" does not exists')
        result = []
        for row in self.data_rows:
            result.append(row[col])
        return result

    def append_col(self, col: str, default=''):
        if col in self.fields:
            raise ValueError('Duplicated col')
        self.fields.append(col)
        for row in self.data_rows:
            row[col] = default

    def print_row(self, index: int):
        row = self.data_rows[index]
        result = ''
        for k in row:
            result += f'{k}: {row[k].value}, '
        return result

    def set_row_style(self, row: Union[Row, int], style: Style) -> None:
        if isinstance(row, int):
            row = self.data_rows[row]

        for c in row:
            row[c].style = style

    def remove_row(self, row: Row) -> None:
        self.data_rows.remove(row)

    def import_json(self, path: str) -> None:
        with open(path, 'r') as f:
            data = json.load(f)
        if not isinstance(data, list):
            raise ValueError('invalid file format')
        for row in data:
            self.append_row(row)

    def to_csv(self, out: str = '') -> None:
        if out == '':
            out = self.name + '.csv'

        with open(out, 'w') as f:
            w = csv.DictWriter(f, self.fields)
            v = {}
            for r in self.data_rows:
                for key in r:
                    v[key] = r[key].value
                w.writerow(v)

    def to_json(self, out: str = '') -> None:
        if out == '':
            out = self.name + '.json'
        data = []
        for r in self.data_rows:
            v = {}
            for k in r:
                v[k] = r[k].value
            data.append(v)
        with open(out, 'w') as f:
            json.dump(data, f)

    # def split_rows(path: str, row_count: int, name_by: str):
    #     filenames = {}

    def sort_by(self, by: str, desc=False):
        copied: List[Row] = [*self.data_rows]
        result: List[Row] = []
        for i in range(len(self.data_rows)):
            min = copied[0]
            for j in range(len(copied)):
                if desc:
                    if copied[j][by].value > min[by].value:
                        min = copied[j]
                else:
                    if copied[j][by].value < min[by].value:
                        min = copied[j]
            copied.remove(min)
            result.append(min)
        self.data_rows = result

    def beautify(self, by: str) -> List[Row]:
        if isinstance(by, str):
            grouped = []
            ungrouped = copy(self.data_rows)
            while ungrouped.__len__() > 0:
                counter = 0
                current = ungrouped[0][by].value
                while counter < ungrouped.__len__():
                    if ungrouped[counter][by].value == current:
                        grouped.append(ungrouped[counter])
                        ungrouped.remove(ungrouped[counter])
                        counter -= 1

                    counter += 1
            return grouped

    def __eq__(self, other):
        if self.name == other.name:
            return True
        return False


class Dataset:

    def __init__(self, path: str, suppress_warning=False):
        self.workbook = load_workbook(path)
        self.sheets = []
        self.filename = os.path.basename(path)
        self.backup_name = self.filename + '.bak'
        self.path = os.path.dirname(path)
        self.suppress_warning = suppress_warning
        sheet: Worksheet
        for sheet in self.workbook.worksheets:
            if sheet.max_row == 0:
                raise EmptySheetException
            self.sheets.append(Sheet(self.suppress_warning, sheet))
        # self.workbook.close()

    def _resolve_cell_notation(self, s: str) -> Tuple[int, int]:
        """
        convert cell notation to (row, col) tuple
        :param s: Cell Notation
        :return: (row, col)
        """
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        col: str = ''
        row: Union[str, int]
        i = 0
        for i in range(len(s)):
            if s[i].isalpha():
                col += s[i]
        row = s[i:]
        row = int(row)
        real_col = 0
        # resolve col
        for i in range(len(col)):
            real_col += 26 ** (len(col) - 1 - i) * (letters.index(col[i]) + 1)

        return row - 1, real_col - 1

    def get_sheet_by_index(self, index: int) -> Sheet:
        return self.sheets[index]

    def get_sheet_by_name(self, name: str) -> Union[Sheet, None]:
        t: Sheet
        for t in self.sheets:
            if t.name.lower() == name.lower():
                return t
        else:
            return None

    def does_exist(self, name: str) -> bool:
        for t in self.sheets:
            if t.name == name:
                return True
        else:
            return False

    def duplicate(self, path: str, headers_only: bool = False):
        file = open_file(path)
        for sheet in self.sheets:
            sh = sheet.duplicate(sheet.name, headers_only)
            file.sheets.append(sh)
        return file

    def filter(self, table: Sheet, callback: Callable[[Row], Union[None, bool]]) -> List[Row]:
        return table.filter(callback)

    def find(self, sheet: Sheet, **kwargs) -> List[Row]:
        result = sheet.find(**kwargs)

        return result

    def append_row(self, sheet: Union[Sheet, str], content: Row) -> None:
        if isinstance(sheet, str):
            sheet = self.get_sheet_by_name(sheet)
        if sheet is None:
            raise NameError(f'{sheet} does not exist')
        sheet.append_row(content)

    def add_sheet(self, name: str, fields: List[str]) -> Sheet:
        if self.does_exist(name):
            raise Exception('Sheet already exists')
        table = Sheet(self.suppress_warning, name)
        table.fields = fields
        self.sheets.append(table)
        return table

    def create_sheet_by_json(self, name: str, data: Union[str, list, dict]) -> Sheet:
        if isinstance(data, str):
            with open(data, 'r') as f:
                data: Union[list, dict] = json.load(f)

        if isinstance(data, list):
            header = data[0].keys()
        elif isinstance(data, dict):
            header = data.keys()
        else:
            raise ValueError('corrupted file')
        sheet = self.add_sheet(name, header)
        if isinstance(data, list):
            for d in data:
                sheet.append_row(d)
        return sheet

    def import_json(self, path: str) -> None:
        with open(path, 'r') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            raise ValueError('invalid format')

        for key in data:
            self.create_sheet_by_json(key, data[key])

    def export_json(self, out: str):
        json_sheets = {}
        for sheet in self.sheets:
            data = []
            for r in sheet.data_rows:
                v = {}
                for k in r:
                    v[k] = r[k].value
                data.append(v)
            json_sheets[sheet.name] = data
        with open(out, 'w') as f:
            json.dump(json_sheets, f)

    def to_sqlite(self, out: str):
        conn = sqlite3.connect(out)
        cur = conn.cursor()
        for sheet in self.sheets:
            current_table = sheet.name
            cmd = f"CREATE TABLE '{current_table}' ({','.join(sheet.fields)})"
            cur.execute(cmd)
            conn.commit()
            for row in sheet.data_rows:
                values = ''
                for cell in row.values():
                    if isinstance(cell.value, float):
                        values += str(cell.value)
                    else:
                        values += '"' + cell.value + '"'
                    values += ','
                values = values[: -1]
                cmd = f"INSERT INTO {current_table} VALUES ({values})"
                cur.execute(cmd)
        conn.commit()
        conn.close()

    def merge_file(self, path: str, force: bool = False) -> None:
        workbook = load_workbook(path, read_only=True)
        sheet: Worksheet
        for sheet in workbook.worksheets:
            tbl = self.get_sheet_by_name(sheet.title)
            if tbl is not None:
                if force:
                    headers_to_merge = sheet.rows.__next__()
                    for i in range(len(headers_to_merge)):
                        headers_to_merge[i] = headers_to_merge[i].value
                    for h in headers_to_merge:
                        if h in tbl.fields:
                            headers_to_merge.remove(h)
                    tbl.fields.extend(headers_to_merge)
                self._merge_table(sheet, tbl)
            else:
                tbl = Sheet(self.suppress_warning, sheet.title)
                try:
                    headers = sheet.rows.__next__()
                except IndexError:
                    raise ValueError('File has no headers')
                for h in headers:
                    tbl.fields.append(h.value)
                self._merge_table(sheet, tbl)
                self.sheets.append(tbl)

    def _merge_table(self, sheet: Worksheet, new_sheet):
        flg_first_row = True
        for row in sheet.rows:
            # Skip header
            if flg_first_row:
                flg_first_row = False
                continue

            new_row = []
            for cell in row:
                new_row.append(cell.value)
            new_sheet.append_row(new_row)

    def split_sheets_to_file(self):
        for s in self.sheets:
            doc = open_file(s.name + '.xlsx')
            doc.add_sheet(s.name, s.fields)
            for row in s.data_rows:
                doc.append_row(s.name, row)
            doc.save(backup=False)

    def remove_sheet(self, sheet: Sheet) -> None:
        self.sheets.remove(sheet)

    def remove_sheet_by_index(self, index: int):
        pass

    def save(self, *, backup=True, row_height=0, col_width=0):
        # make backup & delete
        if os.path.exists(os.path.join(self.path, self.filename)) and backup:
            shutil.copy(os.path.join(self.path, self.filename), os.path.join(self.path, self.backup_name))
            os.remove(os.path.join(self.path, self.filename))

        # open new file
        filename = os.path.join(self.path, self.filename)
        workbook = xlsxwriter.Workbook(filename, {'default_date_format':
                                                      'yyyy/mm/dd'})
        for table in self.sheets:
            sheet: xlsxwriter.workbook.Worksheet = workbook.add_worksheet(table.name)
            merged = []
            pointer = Pointer(table.header_starts, 0)
            # write useless file head
            for i in range(table.header_starts):
                for cell in table.file_head[i]:
                    if isinstance(cell, MergedCell):
                        if cell in merged:
                            continue
                        sheet.merge_range(cell.start_row,
                                          cell.start_col,
                                          cell.end_row,
                                          cell.end_col if cell.end_col != -1 else table.fields.__len__() - 1,
                                          data=cell.value,
                                          cell_format=workbook.add_format(Style('center', 'center').attr()))
                        merged.append(cell)
                    else:
                        sheet.write(cell.value)

            # write headers
            for field in table.fields:
                sheet.write(pointer.row, pointer.col, field, workbook.add_format(table.header_style.attr()))
                pointer.next_col()
            pointer.next_row()
            for data_row in table.data_rows:
                for col in table.fields:
                    data = data_row[col]
                    if isinstance(data.value, datetime.date) \
                            or isinstance(data.value, datetime.time) \
                            or isinstance(data.value, datetime.datetime):

                        if isinstance(data.value, datetime.date):
                            sheet.write(pointer.row, pointer.col, str(data.value.isoformat()),
                                        workbook.add_format(data.attr()))
                        elif isinstance(data.value, datetime.time):
                            sheet.write(pointer.row, pointer.col, str(data.value.isoformat()),
                                        workbook.add_format(data.attr()))
                        else:
                            sheet.write(pointer.row, pointer.col, str(data.value), workbook.add_format(data.attr()))

                    else:

                        if isinstance(data, ImageCell):
                            if isinstance(data.data, str):
                                sheet.insert_image(pointer.row, pointer.col, data.data,
                                                   {'y_offset': 10, 'x_offset': 10})
                            else:
                                data.data.seek(0)
                                sheet.insert_image(pointer.row, pointer.col, data.value,
                                                   {'image_data': data.data, 'y_offset': 10, 'x_offset': 10})
                            img: Image.Image = Image.open(data.data)
                            width, height = img.size
                            if row_height == 0:
                                sheet.set_row(pointer.row, height)
                            else:
                                sheet.set_row(pointer.row, row_height)
                            if col_width != 0:
                                sheet.set_column(pointer.col, pointer.col, (col_width / 8))
                            else:
                                sheet.set_column(pointer.col, pointer.col, (width / 8))

                        elif isinstance(data, FormulaCell):
                            sheet.write_formula(pointer.row, pointer.col, data.formula,
                                                workbook.add_format(data.attr()))
                        else:
                            sheet.write(pointer.row, pointer.col, data.value, workbook.add_format(data.attr()))
                    pointer.next_col()
                pointer.next_row()
        workbook.close()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()

    def __enter__(self):
        return self


def open_file(path: str, suppress_warning=False) -> Dataset:
    return Dataset(path, suppress_warning=suppress_warning)
